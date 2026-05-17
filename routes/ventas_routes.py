import logging
import io
from datetime import date, datetime
from flask import Blueprint, render_template, jsonify, session, request, send_file
from config import MAX_FILAS_EXPORTAR
from openpyxl import Workbook
from services.excel_helpers import (
    C, xl_cell, xl_row_bg, fmt_dt, estado_venta,
    xl_titulo_hoja, xl_fila_headers, xl_fila_totales,
    xl_col_widths, xl_badge_estado, send_excel
)

from services.ventas_service import (
    listar_ventas_listas_service,
    registrar_pago_final_service,
    listar_entregas_pendientes_service,
    guardar_venta_service,
    marcar_lista_service,
    marcar_entregada,
    obtener_venta,
    obtener_detalles_venta,
    eliminar_venta_service,
    historial_ventas_service,
)

from middleware.auth_middleware import admin_required
from extensions import limiter

logger = logging.getLogger(__name__)

ventas_bp = Blueprint("ventas", __name__)


@ventas_bp.route("/ventas/guardar", methods=["POST"])
@limiter.limit("30 per minute")
def guardar_venta():
    id_usuario = session.get("id_usuario")
    if not id_usuario:
        return jsonify({"ok": False, "error": "Sesión expirada. Vuelve a iniciar sesión."}), 401
    try:
        id_venta = guardar_venta_service(request.form, id_usuario)
        return jsonify({"ok": True, "id_venta": id_venta}), 200
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception:
        logger.exception("Error en guardar_venta id_usuario=%s", id_usuario)
        return jsonify({"ok": False, "error": "Error interno del servidor"}), 500


@ventas_bp.route("/ventas")
def ventas():
    return render_template("ventas_crear.html")


@ventas_bp.route("/ventas/entregar/<int:id_venta>", methods=["POST"])
@limiter.limit("30 per minute")
def entregar_venta(id_venta):
    id_usuario = session.get("id_usuario")

    try:
        if marcar_entregada(id_venta, id_usuario):
            return jsonify({
                "ok": True,
                "message": "Venta entregada correctamente"
            })
        else:
            return jsonify({
                "ok": False,
                "error": "La venta ya fue entregada o no existe"
            })

    except Exception:
        logger.exception("Error en entregar_venta id_venta=%s id_usuario=%s", id_venta, id_usuario)
        return jsonify({"ok": False, "error": "Error interno del servidor"}), 500


@ventas_bp.route("/ventas/ticket/<int:id_venta>")
def venta_ticket(id_venta):
    venta = obtener_venta(id_venta)
    if not venta:
        return render_template("404.html"), 404

    detalles_dict = obtener_detalles_venta([id_venta])
    detalles = detalles_dict.get(id_venta, [])

    return render_template(
        "ticket_venta.html",
        venta=venta,
        detalles=detalles
    )

@ventas_bp.route("/ventas/listas")
def ventas_listas():
    id_negocio = request.args.get("id_negocio", type=int)

    data = listar_ventas_listas_service(id_negocio)

    return render_template(
        "ventas_listas.html",
        **data
    )


@ventas_bp.route("/ventas/pendientes")
def ventas_pendientes():
    try:
        id_negocio = request.args.get("id_negocio", type=int)

        data = listar_entregas_pendientes_service(id_negocio)

        return render_template(
            "ventas_pendientes.html",
            **data
        )

    except Exception:
        logger.exception("Error en ventas_pendientes id_negocio=%s", id_negocio)
        return render_template("403.html"), 500


@ventas_bp.route("/ventas/marcar-lista/<int:id_venta>", methods=["POST"])
@limiter.limit("30 per minute")
def marcar_lista(id_venta):
    try:
        id_usuario = session.get('id_usuario')
        if marcar_lista_service(id_venta, id_usuario):
            return jsonify({
                "ok": True,
                "message": "Venta marcada como lista correctamente"
            })
        else:
            return jsonify({
                "ok": False,
                "error": "La venta ya está lista o fue entregada"
            })
    except Exception:
        logger.exception("Error en marcar_lista id_venta=%s id_usuario=%s", id_venta, id_usuario)
        return jsonify({"ok": False, "error": "Error interno del servidor"}), 500


@ventas_bp.route("/ventas/detalles/<int:id_venta>")
def detalles_venta(id_venta):
    detalles = obtener_detalles_venta([id_venta])  
    return jsonify(detalles.get(id_venta, []))

@ventas_bp.route("/ventas/pago-final", methods=["POST"])
@limiter.limit("30 per minute")
def registrar_pago_final():
    data = request.get_json(silent=True) or {}
    id_usuario = session.get("id_usuario")

    try:
        mensaje = registrar_pago_final_service(data, id_usuario)
        return jsonify({"ok": True, "message": mensaje})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception:
        logger.exception("Error en registrar_pago_final id_usuario=%s", id_usuario)
        return jsonify({"ok": False, "error": "Error interno del servidor"}), 500


@ventas_bp.route("/ventas/eliminar/<int:id_venta>", methods=["POST"])
@admin_required
@limiter.limit("30 per minute")
def eliminar_venta_route(id_venta):
    id_usuario = session.get("id_usuario")

    try:
        eliminar_venta_service(id_venta, id_usuario)
        return jsonify({"ok": True, "message": "Venta eliminada correctamente"})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception:
        logger.exception("Error en eliminar_venta id_venta=%s id_usuario=%s", id_venta, id_usuario)
        return jsonify({"ok": False, "error": "Error interno del servidor"}), 500

@ventas_bp.route("/ventas/historial")
@admin_required
def historial_ventas():

    id_negocio   = request.args.get("id_negocio",  type=int)
    fecha_inicio = request.args.get("fecha_inicio") or None
    fecha_fin    = request.args.get("fecha_fin")    or None
    pagina       = request.args.get("pagina", 1,    type=int)
    q            = request.args.get("q", "").strip() or None

    mostrar_eliminadas = request.args.get('eliminadas') == '1'
    data = historial_ventas_service(id_negocio, fecha_inicio, fecha_fin, pagina, mostrar_eliminadas, q=q)

    return render_template("historial_ventas.html", **data)



@ventas_bp.route("/ventas/historial/exportar")
@admin_required
def exportar_historial_excel():
    from ventas import obtener_historial_ventas
    from pagos import obtener_pagos_venta

    id_negocio   = request.args.get("id_negocio",  type=int)
    fecha_inicio = request.args.get("fecha_inicio") or None
    fecha_fin    = request.args.get("fecha_fin")    or None

    ventas = obtener_historial_ventas(id_negocio, fecha_inicio, fecha_fin, limit=MAX_FILAS_EXPORTAR, offset=0)
    ids_venta    = [v["id_venta"] for v in ventas]
    detalles_map = obtener_detalles_venta(ids_venta)
    pagos_map    = obtener_pagos_venta(ids_venta)

    filtro_txt = "  ".join(filter(None, [
        f"Negocio ID: {id_negocio}" if id_negocio   else "",
        f"Desde: {fecha_inicio}"    if fecha_inicio  else "",
        f"Hasta: {fecha_fin}"       if fecha_fin     else "",
    ])) or "Sin filtros — todos los registros"

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Resumen ventas"
    ws1.freeze_panes = "A4"
    COLS1 = ["# Recibo","Negocio","Cliente","Teléfono","Fecha recibo","Fecha estimada",
             "Fecha lista","Fecha entrega","Total ($)","Cobrado ($)","Saldo ($)",
             "Estado","Registrada por","Entregada por"]
    xl_titulo_hoja(ws1, "Historial de Ventas", len(COLS1), filtro_txt)
    xl_fila_headers(ws1, COLS1)
    for i, v in enumerate(ventas):
        r=i+4; bg=xl_row_bg(i)
        estado=estado_venta(v); total=float(v.get("total") or 0)
        cobrado=float(v.get("total_pagado") or 0); saldo=max(total-cobrado,0)
        xl_cell(ws1,r,1, f"#{v['id_venta']}",fg=bg,align="center",bold=True)
        xl_cell(ws1,r,2, v.get("negocio",""),fg=bg)
        xl_cell(ws1,r,3, f"{v['nombre']} {v['apellido']}",fg=bg)
        xl_cell(ws1,r,4, v.get("telefono") or "—",fg=bg)
        xl_cell(ws1,r,5, fmt_dt(v.get("fecha_recibo")),fg=bg)
        xl_cell(ws1,r,6, fmt_dt(v.get("fecha_estimada")),fg=bg)
        xl_cell(ws1,r,7, fmt_dt(v.get("fecha_lista")),fg=bg)
        xl_cell(ws1,r,8, fmt_dt(v.get("fecha_entrega")),fg=bg)
        xl_cell(ws1,r,9, total,  fg=bg,bold=True,align="right",num_fmt='"$"#,##0.00')
        xl_cell(ws1,r,10,cobrado,fg=bg,bold=True,align="right",num_fmt='"$"#,##0.00')
        xl_cell(ws1,r,11,saldo,  fg=bg,bold=True,align="right",num_fmt='"$"#,##0.00',color=C["rojo"] if saldo>0 else C["verde"])
        xl_badge_estado(ws1,r,12,estado)
        xl_cell(ws1,r,13,v.get("usuario_creo") or "—",fg=bg)
        xl_cell(ws1,r,14,v.get("usuario_entrego") or "—",fg=bg)
        ws1.row_dimensions[r].height=16
    if ventas: xl_fila_totales(ws1,len(ventas)+4,len(COLS1),[9,10,11])
    xl_col_widths(ws1,[10,16,26,16,20,20,18,18,13,13,13,13,18,18])

    ws2 = wb.create_sheet("Artículos")
    ws2.freeze_panes = "A4"
    COLS2 = ["# Recibo","Negocio","Cliente","Tipo artículo","Descripción","Material / Notas","Cantidad","Servicio","Precio servicio ($)","Comentario"]
    xl_titulo_hoja(ws2,"Artículos por Venta",len(COLS2),filtro_txt)
    xl_fila_headers(ws2,COLS2)
    r2=4
    for v in ventas:
        detalles=detalles_map.get(v["id_venta"],[])
        if not detalles:
            bg=xl_row_bg(r2)
            xl_cell(ws2,r2,1,f"#{v['id_venta']}",fg=bg,bold=True,align="center")
            xl_cell(ws2,r2,2,v.get("negocio",""),fg=bg)
            xl_cell(ws2,r2,3,f"{v['nombre']} {v['apellido']}",fg=bg)
            for ci in range(4,11): xl_cell(ws2,r2,ci,"—",fg=bg,color=C["gris_txt"],align="center")
            ws2.row_dimensions[r2].height=15; r2+=1; continue
        for det in detalles:
            tipo=det.get("tipo_articulo",""); datos=det.get("datos") or {}; coment=det.get("comentario") or ""; servicios=det.get("servicios",[])
            if tipo=="calzado":   desc=f"{datos.get('tipo','')} {datos.get('marca','')}".strip(); mat=f"Color: {datos.get('color_base','—')}  Material: {datos.get('material','—')}"; cant=1
            elif tipo=="confeccion": desc=f"{datos.get('tipo','')} {datos.get('marca','')}".strip(); mat=f"Material: {datos.get('material','—')}"; cant=datos.get("cantidad",1)
            elif tipo=="maquila": desc=datos.get("tipo","—"); mat=f"Precio unitario: ${float(datos.get('precio_unitario') or 0):.2f}"; cant=datos.get("cantidad",1)
            else: desc=mat="—"; cant="—"
            if not servicios:
                bg=xl_row_bg(r2)
                xl_cell(ws2,r2,1,f"#{v['id_venta']}",fg=bg,bold=True,align="center"); xl_cell(ws2,r2,2,v.get("negocio",""),fg=bg); xl_cell(ws2,r2,3,f"{v['nombre']} {v['apellido']}",fg=bg)
                xl_cell(ws2,r2,4,tipo.capitalize(),fg=bg,align="center"); xl_cell(ws2,r2,5,desc,fg=bg); xl_cell(ws2,r2,6,mat,fg=bg,wrap=True)
                xl_cell(ws2,r2,7,cant,fg=bg,align="center"); xl_cell(ws2,r2,8,"—",fg=bg,color=C["gris_txt"],align="center"); xl_cell(ws2,r2,9,"—",fg=bg,color=C["gris_txt"],align="center")
                xl_cell(ws2,r2,10,coment,fg=bg,wrap=True,italic=bool(coment)); ws2.row_dimensions[r2].height=15; r2+=1
            else:
                for si,svc in enumerate(servicios):
                    bg=xl_row_bg(r2)
                    xl_cell(ws2,r2,1,f"#{v['id_venta']}" if si==0 else "",fg=bg,bold=si==0,align="center"); xl_cell(ws2,r2,2,v.get("negocio","") if si==0 else "",fg=bg)
                    xl_cell(ws2,r2,3,f"{v['nombre']} {v['apellido']}" if si==0 else "",fg=bg); xl_cell(ws2,r2,4,tipo.capitalize() if si==0 else "",fg=bg,align="center")
                    xl_cell(ws2,r2,5,desc if si==0 else "",fg=bg); xl_cell(ws2,r2,6,mat if si==0 else "",fg=bg,wrap=True); xl_cell(ws2,r2,7,cant if si==0 else "",fg=bg,align="center")
                    xl_cell(ws2,r2,8,svc.get("nombre",""),fg=bg); xl_cell(ws2,r2,9,float(svc.get("precio_aplicado") or 0),fg=bg,align="right",num_fmt='"$"#,##0.00')
                    xl_cell(ws2,r2,10,coment if si==0 else "",fg=bg,wrap=True,italic=bool(coment)); ws2.row_dimensions[r2].height=15; r2+=1
    xl_col_widths(ws2,[10,16,26,14,24,28,10,24,18,24])

    ws3 = wb.create_sheet("Pagos")
    ws3.freeze_panes = "A4"
    COLS3 = ["# Recibo","Negocio","Cliente","Tipo pago","Método","Monto ($)","Total venta ($)","Registrada por","Entregada por"]
    xl_titulo_hoja(ws3,"Pagos por Venta",len(COLS3),filtro_txt)
    xl_fila_headers(ws3,COLS3)
    r3=4
    for v in ventas:
        pagos=pagos_map.get(v["id_venta"],[]); total=float(v.get("total") or 0)
        if not pagos:
            bg=xl_row_bg(r3)
            xl_cell(ws3,r3,1,f"#{v['id_venta']}",fg=bg,bold=True,align="center"); xl_cell(ws3,r3,2,v.get("negocio",""),fg=bg); xl_cell(ws3,r3,3,f"{v['nombre']} {v['apellido']}",fg=bg)
            xl_cell(ws3,r3,4,"Sin pagos",fg=bg,color=C["gris_txt"],italic=True); xl_cell(ws3,r3,5,"—",fg=bg,color=C["gris_txt"],align="center"); xl_cell(ws3,r3,6,"—",fg=bg,color=C["gris_txt"],align="center")
            xl_cell(ws3,r3,7,total,fg=bg,align="right",num_fmt='"$"#,##0.00'); xl_cell(ws3,r3,8,v.get("usuario_creo") or "—",fg=bg); xl_cell(ws3,r3,9,v.get("usuario_entrego") or "—",fg=bg)
            ws3.row_dimensions[r3].height=15; r3+=1; continue
        for pi,p in enumerate(pagos):
            bg=xl_row_bg(r3)
            xl_cell(ws3,r3,1,f"#{v['id_venta']}" if pi==0 else "",fg=bg,bold=pi==0,align="center"); xl_cell(ws3,r3,2,v.get("negocio","") if pi==0 else "",fg=bg); xl_cell(ws3,r3,3,f"{v['nombre']} {v['apellido']}" if pi==0 else "",fg=bg)
            xl_cell(ws3,r3,4,(p.get("tipo_pago_venta") or "—").capitalize(),fg=bg); xl_cell(ws3,r3,5,(p.get("tipo_pago") or "—").capitalize(),fg=bg)
            xl_cell(ws3,r3,6,float(p.get("monto") or 0),fg=bg,bold=True,align="right",color=C["verde"],num_fmt='"$"#,##0.00')
            xl_cell(ws3,r3,7,total if pi==0 else "",fg=bg,align="right",num_fmt='"$"#,##0.00')
            xl_cell(ws3,r3,8,v.get("usuario_creo") or "—" if pi==0 else "",fg=bg); xl_cell(ws3,r3,9,v.get("usuario_entrego") or "—" if pi==0 else "",fg=bg)
            ws3.row_dimensions[r3].height=15; r3+=1
    xl_col_widths(ws3,[10,16,26,16,16,14,14,18,18])

    return send_excel(wb, "historial_ventas")

@ventas_bp.route("/ventas/<int:id_venta>/historial")
def historial_venta_por_id(id_venta):
    from ventas import obtener_historial_venta
    data = obtener_historial_venta(id_venta)
    # Serializar fechas
    from datetime import datetime, date
    result = []
    for row in data:
        r = dict(row)
        for k, v in r.items():
            if isinstance(v, (datetime, date)):
                r[k] = v.isoformat()
        result.append(r)
    return jsonify(result)