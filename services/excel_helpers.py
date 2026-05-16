import io
from datetime import datetime, date
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


C = {
    "azul":        "1E7FD6",
    "azul_cl":     "E6F3FF",
    "azul_med":    "BFDBFE",
    "blanco":      "FFFFFF",
    "gris":        "F8FBFF",
    "verde":       "22C55E",
    "verde_cl":    "DCFCE7",
    "amarillo":    "F59E0B",
    "amarillo_cl": "FEF9C3",
    "rojo":        "E53935",
    "rojo_cl":     "FEE2E2",
    "gris_txt":    "6B7280",
    "dark":        "1F2937",
}

COLOR_ESTADO    = {"Entregada": C["verde"],    "Lista": C["amarillo"],    "Pendiente": C["rojo"]}
COLOR_ESTADO_CL = {"Entregada": C["verde_cl"], "Lista": C["amarillo_cl"], "Pendiente": C["rojo_cl"]}

BORDE = Border(
    left=Side(style="thin", color="CFD8E3"),
    right=Side(style="thin", color="CFD8E3"),
    top=Side(style="thin", color="CFD8E3"),
    bottom=Side(style="thin", color="CFD8E3"),
)

BORDE_MED = Border(
    left=Side(style="medium", color="1E7FD6"),
    right=Side(style="medium", color="1E7FD6"),
    top=Side(style="medium", color="1E7FD6"),
    bottom=Side(style="medium", color="1E7FD6"),
)


def xl_cell(ws, r, col, value="", bold=False, fg=None, color=None,
            align="left", num_fmt=None, italic=False, size=9, wrap=False):
    """Escribe y estiliza una celda."""
    c = ws.cell(row=r, column=col, value=value)
    c.font = Font(
        bold=bold, italic=italic,
        color=color or C["dark"],
        name="Arial", size=size
    )
    if fg:
        c.fill = PatternFill("solid", fgColor=fg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = BORDE
    if num_fmt:
        c.number_format = num_fmt
    return c


def xl_head(ws, r, col, value, size=9):
    return xl_cell(ws, r, col, value,
                   bold=True, fg=C["azul"], color=C["blanco"],
                   align="center", size=size)


def xl_subhead(ws, r, col, value):
    return xl_cell(ws, r, col, value,
                   bold=True, fg=C["azul_cl"], color=C["azul"], align="left")


def xl_row_bg(i):
    return C["gris"] if i % 2 == 0 else C["blanco"]


def xl_titulo_hoja(ws, texto, ncols, subtexto=""):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value     = texto
    t.font      = Font(bold=True, color=C["blanco"], name="Arial", size=13)
    t.fill      = PatternFill("solid", fgColor=C["azul"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    s = ws["A2"]
    s.value     = subtexto or "Fresh Steps"
    s.font      = Font(italic=True, color=C["gris_txt"], name="Arial", size=9)
    s.fill      = PatternFill("solid", fgColor=C["azul_cl"])
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16


def xl_fila_headers(ws, headers, fila=3, altura=22):
    for ci, h in enumerate(headers, 1):
        xl_head(ws, fila, ci, h)
    ws.row_dimensions[fila].height = altura


def xl_fila_totales(ws, fila, ncols, cols_suma):
    fd = 4
    ld = fila - 1

    ws.merge_cells(f"A{fila}:{get_column_letter(cols_suma[0] - 1)}{fila}")
    ct = ws.cell(row=fila, column=1, value="TOTALES")
    ct.font      = Font(bold=True, color=C["blanco"], name="Arial", size=10)
    ct.fill      = PatternFill("solid", fgColor=C["azul"])
    ct.alignment = Alignment(horizontal="right", vertical="center")
    ct.border    = BORDE

    for col in cols_suma:
        L  = get_column_letter(col)
        tc = ws.cell(row=fila, column=col, value=f"=SUM({L}{fd}:{L}{ld})")
        tc.font         = Font(bold=True, color=C["blanco"], name="Arial", size=10)
        tc.fill         = PatternFill("solid", fgColor=C["azul"])
        tc.alignment    = Alignment(horizontal="right", vertical="center")
        tc.number_format = '"$"#,##0.00'
        tc.border        = BORDE

    for col in range(1, ncols + 1):
        if col not in cols_suma and col != 1:
            ec = ws.cell(row=fila, column=col)
            if not ec.fill or ec.fill.fgColor.rgb == "00000000":
                ec.fill   = PatternFill("solid", fgColor=C["azul"])
                ec.border = BORDE

    ws.row_dimensions[fila].height = 20


def xl_col_widths(ws, widths):
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def fmt_dt(dt):
    if not dt:
        return "—"
    if isinstance(dt, datetime):
        return dt.strftime("%d/%m/%Y %H:%M")
    if isinstance(dt, date):
        return dt.strftime("%d/%m/%Y")
    return str(dt)


def estado_venta(v):
    if v.get("fecha_entrega"):
        return "Entregada"
    if v.get("fecha_lista"):
        return "Lista"
    return "Pendiente"


def xl_badge_estado(ws, r, col, estado, borde=None):
    ce = ws.cell(row=r, column=col, value=estado)
    ce.font      = Font(bold=True, color=COLOR_ESTADO[estado], name="Arial", size=9)
    ce.fill      = PatternFill("solid", fgColor=COLOR_ESTADO_CL[estado])
    ce.alignment = Alignment(horizontal="center", vertical="center")
    ce.border    = borde or BORDE


def xl_badge_activo(ws, r, col, activo, borde=None):
    valor = "Activo" if activo else "Eliminado"
    color = C["verde"] if activo else C["rojo"]
    fondo = C["verde_cl"] if activo else C["rojo_cl"]
    ce = ws.cell(row=r, column=col, value=valor)
    ce.font      = Font(bold=True, color=color, name="Arial", size=9)
    ce.fill      = PatternFill("solid", fgColor=fondo)
    ce.alignment = Alignment(horizontal="center", vertical="center")
    ce.border    = borde or BORDE


def _build_ws_resumen(ws, pedidos, pagos_map, nombre_cliente, filtro_txt):
    ws.freeze_panes = "A4"
    COLS = ["# Recibo","Negocio","Fecha recibo","Fecha estimada","Fecha lista","Fecha entrega",
            "Total ($)","Cobrado ($)","Saldo ($)","Estado"]
    xl_titulo_hoja(ws, f"Historial de pedidos — {nombre_cliente}", len(COLS), filtro_txt)
    xl_fila_headers(ws, COLS)
    for i, p in enumerate(pedidos):
        r = i + 4; bg = xl_row_bg(i)
        estado  = estado_venta(p)
        total   = float(p.get("total") or 0)
        pagos   = pagos_map.get(p["id_venta"], [])
        cobrado = sum(float(pg["monto"]) for pg in pagos)
        saldo   = max(total - cobrado, 0)
        xl_cell(ws, r, 1,  f"#{p['id_venta']}",           fg=bg, bold=True, align="center")
        xl_cell(ws, r, 2,  p.get("negocio",""),            fg=bg)
        xl_cell(ws, r, 3,  fmt_dt(p.get("fecha_recibo")),  fg=bg)
        xl_cell(ws, r, 4,  fmt_dt(p.get("fecha_estimada")),fg=bg)
        xl_cell(ws, r, 5,  fmt_dt(p.get("fecha_lista")),   fg=bg)
        xl_cell(ws, r, 6,  fmt_dt(p.get("fecha_entrega")), fg=bg)
        xl_cell(ws, r, 7,  total,   fg=bg, bold=True, align="right", num_fmt='"$"#,##0.00')
        xl_cell(ws, r, 8,  cobrado, fg=bg, bold=True, align="right", num_fmt='"$"#,##0.00')
        xl_cell(ws, r, 9,  saldo,   fg=bg, bold=True, align="right", num_fmt='"$"#,##0.00',
                color=C["rojo"] if saldo > 0 else C["verde"])
        xl_badge_estado(ws, r, 10, estado)
        ws.row_dimensions[r].height = 16
    if pedidos:
        xl_fila_totales(ws, len(pedidos) + 4, len(COLS), [7, 8, 9])
    xl_col_widths(ws, [10, 18, 20, 20, 18, 18, 13, 13, 13, 13])


def _build_ws_articulos(ws, pedidos, detalles_map, nombre_cliente, filtro_txt):
    ws.freeze_panes = "A4"
    COLS = ["# Recibo","Negocio","Tipo artículo","Descripción","Material / Notas",
            "Cantidad","Servicio","Precio ($)","Comentario"]
    xl_titulo_hoja(ws, f"Artículos — {nombre_cliente}", len(COLS), filtro_txt)
    xl_fila_headers(ws, COLS)
    r = 4
    for p in pedidos:
        detalles = detalles_map.get(p["id_venta"], [])
        if not detalles:
            bg = xl_row_bg(r)
            xl_cell(ws, r, 1, f"#{p['id_venta']}", fg=bg, bold=True, align="center")
            xl_cell(ws, r, 2, p.get("negocio",""), fg=bg)
            for ci in range(3, 10):
                xl_cell(ws, r, ci, "—", fg=bg, color=C["gris_txt"], align="center")
            ws.row_dimensions[r].height = 15; r += 1; continue
        for det in detalles:
            tipo     = det.get("tipo_articulo","")
            datos    = det.get("datos") or {}
            coment   = det.get("comentario") or ""
            servicios= det.get("servicios", [])
            if tipo == "calzado":
                desc = f"{datos.get('tipo','')} {datos.get('marca','')}".strip()
                mat  = f"Color: {datos.get('color_base','—')}  Material: {datos.get('material','—')}"
                cant = 1
            elif tipo == "confeccion":
                desc = f"{datos.get('tipo','')} {datos.get('marca','')}".strip()
                mat  = f"Material: {datos.get('material','—')}"
                cant = datos.get("cantidad", 1)
            elif tipo == "maquila":
                desc = datos.get("tipo","—")
                mat  = f"Precio unitario: ${float(datos.get('precio_unitario') or 0):.2f}"
                cant = datos.get("cantidad", 1)
            else:
                desc = mat = "—"; cant = "—"
            for si, svc in enumerate(servicios if servicios else [None]):
                bg = xl_row_bg(r)
                xl_cell(ws,r,1,f"#{p['id_venta']}" if si==0 else "",fg=bg,bold=si==0,align="center")
                xl_cell(ws,r,2,p.get("negocio","") if si==0 else "",fg=bg)
                xl_cell(ws,r,3,tipo.capitalize() if si==0 else "",fg=bg,align="center")
                xl_cell(ws,r,4,desc if si==0 else "",fg=bg)
                xl_cell(ws,r,5,mat if si==0 else "",fg=bg,wrap=True)
                xl_cell(ws,r,6,cant if si==0 else "",fg=bg,align="center")
                xl_cell(ws,r,7,svc.get("nombre","") if svc else "—",fg=bg)
                xl_cell(ws,r,8,float(svc.get("precio_aplicado") or 0) if svc else "—",
                        fg=bg,align="right",num_fmt='"$"#,##0.00' if svc else None)
                xl_cell(ws,r,9,coment if si==0 else "",fg=bg,wrap=True,italic=bool(coment))
                ws.row_dimensions[r].height = 15; r += 1
    xl_col_widths(ws, [10, 18, 14, 24, 28, 10, 24, 13, 24])


def _build_ws_pagos(ws, pedidos, pagos_map, nombre_cliente, filtro_txt):
    ws.freeze_panes = "A4"
    COLS = ["# Recibo","Negocio","Tipo pago","Método","Monto ($)","Total venta ($)"]
    xl_titulo_hoja(ws, f"Pagos — {nombre_cliente}", len(COLS), filtro_txt)
    xl_fila_headers(ws, COLS)
    r = 4
    for p in pedidos:
        pagos = pagos_map.get(p["id_venta"], [])
        total = float(p.get("total") or 0)
        if not pagos:
            bg = xl_row_bg(r)
            xl_cell(ws,r,1,f"#{p['id_venta']}",fg=bg,bold=True,align="center")
            xl_cell(ws,r,2,p.get("negocio",""),fg=bg)
            xl_cell(ws,r,3,"Sin pagos",fg=bg,color=C["gris_txt"],italic=True)
            xl_cell(ws,r,4,"—",fg=bg,align="center")
            xl_cell(ws,r,5,"—",fg=bg,align="center")
            xl_cell(ws,r,6,total,fg=bg,align="right",num_fmt='"$"#,##0.00')
            ws.row_dimensions[r].height = 15; r += 1; continue
        for pi, pg in enumerate(pagos):
            bg = xl_row_bg(r)
            xl_cell(ws,r,1,f"#{p['id_venta']}" if pi==0 else "",fg=bg,bold=pi==0,align="center")
            xl_cell(ws,r,2,p.get("negocio","") if pi==0 else "",fg=bg)
            xl_cell(ws,r,3,(pg.get("tipo_pago_venta") or "—").capitalize(),fg=bg)
            xl_cell(ws,r,4,(pg.get("tipo_pago") or "—").capitalize(),fg=bg)
            xl_cell(ws,r,5,float(pg.get("monto") or 0),fg=bg,bold=True,align="right",
                    color=C["verde"],num_fmt='"$"#,##0.00')
            xl_cell(ws,r,6,total if pi==0 else "",fg=bg,align="right",num_fmt='"$"#,##0.00')
            ws.row_dimensions[r].height = 15; r += 1
    xl_col_widths(ws, [10, 18, 16, 16, 14, 14])


def build_excel_cliente(pedidos, detalles_map, pagos_map, nombre_cliente, filtro_txt):
    wb = Workbook()
    _build_ws_resumen(wb.active, pedidos, pagos_map, nombre_cliente, filtro_txt)
    wb.active.title = "Resumen pedidos"
    ws2 = wb.create_sheet("Artículos")
    _build_ws_articulos(ws2, pedidos, detalles_map, nombre_cliente, filtro_txt)
    ws3 = wb.create_sheet("Pagos")
    _build_ws_pagos(ws3, pedidos, pagos_map, nombre_cliente, filtro_txt)
    return wb


def send_excel(wb, nombre_base):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"{nombre_base}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=nombre,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )